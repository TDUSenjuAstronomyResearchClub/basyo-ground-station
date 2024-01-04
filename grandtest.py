import json
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import messagebox
from threading import Thread
import serial
import matplotlib.pyplot as plt
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg
import folium
from openpyxl.workbook import Workbook
from selenium import webdriver
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from PIL import Image, ImageTk
import io
import os
from datetime import datetime
import pandas as pd

"""COMportを確認し適時変更する"""
port = "COM7"

comand = ("コマンド一覧\ndestination:サンプル採取地点または\nゴール地点の緯度、経度の変更\nfall:機体の落下開始判定\n"
          "landing:機体の着地判定\n//manual:手動制御\n***以降manualで使用***\npicture:写真撮影\nsoil_moisture:土壌水分測定\n"
          "sample:サンプル採取\nw:前進\na:左旋回\nd:右旋回\ns:後退")


class App(tk.Tk):
    def __init__(self):
        super().__init__()

        self.title("Serial Communication")

        # 通信ボタン
        self.button = tk.Button(self, text="Start Communication", command=self.toggle_communication)
        self.button.pack(anchor='ne', padx=10, pady=10)
        self.serial_port = None
        self.is_serial_connected = False
        self.after(100, self.read_serial_data)  # シリアルデータの読み込みを開始

        # tab
        self.notebook = ttk.Notebook(self)
        tab1 = Frame(self)  # main用
        tab2 = Frame(self)  # graph用
        tab3 = Frame(self)  # map用
        self.notebook.add(tab1, text="main")
        self.notebook.add(tab2, text="graph")
        self.notebook.add(tab2, text="graph")
        self.notebook.pack(expand=True, fill="both")

        # 右側コマンド系frame
        self.right = tk.Canvas(tab1, width=280, height=200, borderwidth=0, highlightthickness=0)
        self.command_frame = tk.Frame(self.right)
        self.right.pack(side=tk.RIGHT, fill=tk.BOTH)
        self.right.create_window((0, 0), window=self.command_frame, anchor=tk.NW)

        # 左側データ表示用frame
        self.canvasleft = tk.Canvas(tab1, width=400, height=200, borderwidth=0, highlightthickness=0)
        self.scrollable_lframe = tk.Frame(self.canvasleft)

        self.vsb = tk.Scrollbar(self.scrollable_lframe, orient=tk.VERTICAL, command=self.canvasleft.yview)
        self.canvasleft.configure(yscrollcommand=self.vsb.set)

        self.vsb.pack(side=tk.RIGHT, fill=tk.Y, expand=True)
        self.canvasleft.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.canvasleft.create_window((0, 0), window=self.scrollable_lframe, anchor=tk.NW)

        self.scrollable_lframe.bind("<Configure>", self.on_frame_configure)

        # データ表示用
        self.time_label = tk.Label(self.scrollable_lframe, text="Time:      ", font=("Arial", 11))
        self.time_label.pack(anchor='nw', pady=10)

        self.latitude_label = tk.Label(self.scrollable_lframe, text="Latitude:      ", font=("Arial", 11))
        self.latitude_label.pack(anchor='nw', pady=10)

        self.longitude_label = tk.Label(self.scrollable_lframe, text="Longitude:        ", font=("Arial", 11))
        self.longitude_label.pack(anchor='nw', pady=10)

        self.altitude_label = tk.Label(self.scrollable_lframe, text="Altitude:      ", font=("Arial", 11))
        self.altitude_label.pack(anchor='nw', pady=10)

        self.sample_distance_label = tk.Label(self.scrollable_lframe, text="Sample Distance:        ",
                                              font=("Arial", 11))
        self.sample_distance_label.pack(anchor='nw', pady=10)

        self.sample_azimuth_label = tk.Label(self.scrollable_lframe, text="Sample Azimuth:      ", font=("Arial", 11))
        self.sample_azimuth_label.pack(anchor='nw', pady=10)

        self.goal_distance_label = tk.Label(self.scrollable_lframe, text="Goal Distance:        ", font=("Arial", 11))
        self.goal_distance_label.pack(anchor='nw', pady=10)

        self.goal_azimuth_label = tk.Label(self.scrollable_lframe, text="Goal Azimuth:      ", font=("Arial", 11))
        self.goal_azimuth_label.pack(anchor='nw', pady=10)

        self.acceleration_x_label = tk.Label(self.scrollable_lframe, text="Acceleration X:      ", font=("Arial", 11))
        self.acceleration_x_label.pack(anchor='nw', pady=10)

        self.acceleration_y_label = tk.Label(self.scrollable_lframe, text="Acceleration Y:      ", font=("Arial", 11))
        self.acceleration_y_label.pack(anchor='nw', pady=10)

        self.acceleration_z_label = tk.Label(self.scrollable_lframe, text="Acceleration Z:      ", font=("Arial", 11))
        self.acceleration_z_label.pack(anchor='nw', pady=10)

        self.angular_velocity_x_label = tk.Label(self.scrollable_lframe, text="Angular Velocity X:      ",
                                                 font=("Arial", 11))
        self.angular_velocity_x_label.pack(anchor='nw', pady=10)

        self.angular_velocity_y_label = tk.Label(self.scrollable_lframe, text="Angular Velocity Y:      ",
                                                 font=("Arial", 11))
        self.angular_velocity_y_label.pack(anchor='nw', pady=10)

        self.angular_velocity_z_label = tk.Label(self.scrollable_lframe, text="Angular Velocity Z:      ",
                                                 font=("Arial", 11))
        self.angular_velocity_z_label.pack(anchor='nw', pady=10)

        self.nine_axis_azimuth_label = tk.Label(self.scrollable_lframe, text="Nine Axis Azimuth:        ",
                                                font=("Arial", 11))
        self.angular_velocity_z_label.pack(anchor='nw', pady=10)

        self.temperature_label = tk.Label(self.scrollable_lframe, text="Temperature:        ", font=("Arial", 11))
        self.temperature_label.pack(anchor='nw', pady=10)

        self.humidity_label = tk.Label(self.scrollable_lframe, text="Humidity:      ", font=("Arial", 11))
        self.humidity_label.pack(anchor='nw', pady=10)

        self.pressure_label = tk.Label(self.scrollable_lframe, text="Pressure:      ", font=("Arial", 11))
        self.pressure_label.pack(anchor='nw', pady=10)

        self.battery_label = tk.Label(self.scrollable_lframe, text="Battery:        ", font=("Arial", 11))
        self.battery_label.pack(anchor='nw', pady=10)

        self.distance_label = tk.Label(self.scrollable_lframe, text="Distance:      ", font=("Arial", 11))
        self.distance_label.pack(anchor='nw', pady=10)

        self.soil_label = tk.Label(self.scrollable_lframe, text="soil:      ", font=("Arial", 11))
        self.soil_label.pack(anchor='nw', pady=10)

        # command送信用
        self.sendframe = tk.Frame(self.command_frame)
        self.sendframe.pack(anchor='ne')

        self.send_button = tk.Button(self.sendframe, text="Send Data", command=self.send_data)
        self.send_button.pack(side='right', pady=10, expand=True)
        self.entry = tk.Entry(self.sendframe)
        self.entry.pack(side='right', pady=10, expand=True)

        self.comand_label = tk.Label(self.command_frame, text=comand, font=("Arial", 10))
        self.comand_label.pack(pady=10)

        self.data_text = tk.Entry(self.command_frame, width=40)
        self.data_text.pack(ipady=10, pady=10, expand=True)

        # グラフ
        self.graphframe = tk.Frame(tab2, width=1000, height=800, borderwidth=0, highlightthickness=0)
        fig, (self.acx, self.acy, self.acz, self.avx, self.avy, self.avz, self.tem, self.hum, self.pre,
              self.dis) = plt.subplots(10, 1, sharex='all')
        plt.xlabel('Time')
        self.acx.set_ylabel("Acc X")
        self.acy.set_ylabel("Acc Y")
        self.acz.set_ylabel("Acc Z")
        self.avx.set_ylabel("ang X")
        self.avy.set_ylabel("ang Y")
        self.avz.set_ylabel("ang Z")
        self.tem.set_ylabel("Temperature")
        self.hum.set_ylabel("Humidity")
        self.pre.set_ylabel("Pressure")
        self.dis.set_ylabel("Distance")
        fig.tight_layout()

        """
        self.acx=fig.add_subplot(6,2,1)
        self.acx.set_xlabel("Time")
        self.acx.set_ylabel("Acc X")
        self.acy=fig.add_subplot(6,2,2)
        self.acy.set_xlabel("Time")
        self.acy.set_ylabel("Acc Y")
        self.acz=fig.add_subplot(6,2,3)
        self.acz.set_xlabel("Time")
        self.acz.set_ylabel("Acc Z")
        self.avx=fig.add_subplot(6,2,4)
        self.avx.set_xlabel("Time")
        self.avx.set_ylabel("ang X")
        self.avy=fig.add_subplot(6,2,5)
        self.avy.set_xlabel("Time")
        self.avy.set_ylabel("ang Y")
        self.avz=fig.add_subplot(6,2,6)
        self.avz.set_xlabel("Time")
        self.avz.set_ylabel("ang Z")
        self.tem=fig.add_subplot(6,2,7)
        self.tem.set_xlabel("Time")
        self.tem.set_ylabel("Temperature")
        self.hum=fig.add_subplot(6,2,8)
        self.hum.set_xlabel("Time")
        self.hum.set_ylabel("Humidity")
        self.pre=fig.add_subplot(6,2,9)
        self.pre.set_xlabel("Time")
        self.pre.set_ylabel("Pressure")
        self.bat=fig.add_subplot(6,2,10)
        self.bat.set_xlabel("Time")
        self.bat.set_ylabel("Battery")
        self.dis=fig.add_subplot(6,2,11)
        self.dis.set_xlabel("Time")
        self.dis.set_ylabel("Distance")
        """

        self.fig_canvas = FigureCanvasTkAgg(fig, self.graphframe)
        self.fig_canvas.get_tk_widget().pack(anchor='center', fill=tk.BOTH, expand=True)
        self.graphframe.pack(side=tk.TOP, fill=tk.BOTH, expand=True)
        self.graphframe.bind("<Configure>", self.on_frame_configure)
        # fig.tight_layout()

        # 地図写真用frame
        self.center = tk.Frame(tab1)
        self.center.pack(side=tk.TOP, fill=tk.BOTH)

        # 写真用キャンバス
        # self.cvs = tk.Canvas(self.center, width=500, height=200)
        # self.cvs.pack(expand=True)

        # 地図表示用のフレーム
        self.map_frame = tk.Canvas(self.center, width=700, height=500)
        self.map_frame.pack(expand=True)

        self.protocol("WM_DELETE_WINDOW", self.close)

        # excel処理用
        self.workbook = Workbook()

        # データ
        self.time_data = []
        self.acceleration_x_data = []
        self.acceleration_y_data = []
        self.acceleration_z_data = []
        self.angularvelocity_x_data = []
        self.angularvelocity_y_data = []
        self.angularvelocity_z_data = []
        self.Temperature_data = []
        self.Humidity_data = []
        self.Pressure_data = []
        self.coordinates = []
        self.battery_data = []
        self.distance_data = []

        self.i: int = 2

    def on_frame_configure(self, event):
        self.canvasleft.configure(scrollregion=self.canvasleft.bbox("all"))

    """通信処理関数"""

    def toggle_communication(self):
        if not self.is_serial_connected:
            try:
                self.serial_port = serial.Serial(port, 9600)
                self.is_serial_connected = True
                self.button.configure(text="Stop Communication")
                self.read_serial_data()
            except serial.SerialException as e:
                messagebox.showerror("Error", str(e))
        else:
            self.is_serial_connected = False
            self.serial_port.close()
            self.button.configure(text="Start Communication")

    """シリアル通信動作関数"""

    def read_serial_data(self):
        def read_data():
            self.filename()
            while self.is_serial_connected:
                try:
                    data = self.serial_port.readline().decode('utf-8').strip()
                    json_data = json.loads(data)
                    self.save_to_excel(json_data)
                    self.update_data(json_data)
                except Exception as e:
                    print(str(e))

        Thread(target=read_data, daemon=True).start()

    """ コマンド送信関数 """

    def send_data(self):
        if self.is_serial_connected:
            try:
                data = self.entry.get().strip()
                data = data + '\n'
                encoded_data = data.encode('utf-8')
                self.serial_port.write(encoded_data)
            except Exception as e:
                messagebox.showerror("Error", str(e))
        else:
            messagebox.showerror("Error", "Serial connection is not established.")

    """ 定期通信用データ更新関数 """

    def update_data(self, data):
        img = PhotoImage(file='map.png')
        self.map_frame.create_image(0, 0, anchor='nw', image=img)
        type = data.get("data_type")
        if type == "only_sensor_data":
            self.sensor_data(data)
        elif type == "only_message_data":
            self.text_data(data)
        elif type == "only_picture_data":
            self.picture_data(data)
        elif type == "only_soil_data":
            self.soil_data(data)
        else:
            pass

    """センサデータ表示処理"""

    def sensor_data(self, data):
        time = data.get("time")
        gps = data.get("gps")
        nine_axis = data.get("nine_axis")
        bme280 = data.get("bme280")
        battery = data.get("battery")
        distance = data.get("distance")
        lps25hb = data.get("lps25hb")

        # データ表示
        self.time_label.configure(text=f"Time: {time}")
        self.latitude_label.configure(text=f"Latitude: {gps['latitude']}")
        self.longitude_label.configure(text=f"Longitude: {gps['longitude']}")
        self.altitude_label.configure(text=f"Altitude: {gps['altitude']}")
        self.sample_distance_label.configure(text=f"Sample Distance: {gps['distance']['sample']}")
        self.sample_azimuth_label.configure(text=f"Sample Azimuth: {gps['azimuth']['sample']}")
        self.goal_distance_label.configure(text=f"Goal Distance: {gps['distance']['goal']}")
        self.goal_azimuth_label.configure(text=f"Goal Azimuth: {gps['azimuth']['goal']}")
        self.acceleration_x_label.configure(text=f"Acceleration X: {nine_axis['acceleration']['x']}")
        self.acceleration_y_label.configure(text=f"Acceleration Y: {nine_axis['acceleration']['y']}")
        self.acceleration_z_label.configure(text=f"Acceleration Z: {nine_axis['acceleration']['z']}")
        self.angular_velocity_x_label.configure(text=f"Angular Velocity X: {nine_axis['angular_velocity']['x']}")
        self.angular_velocity_y_label.configure(text=f"Angular Velocity Y: {nine_axis['angular_velocity']['y']}")
        self.angular_velocity_z_label.configure(text=f"Angular Velocity Z: {nine_axis['angular_velocity']['z']}")
        self.nine_axis_azimuth_label.configure(text=f"Nine Axis Azimuth : {nine_axis['azimuth']}")
        self.temperature_label.configure(text=f"Temperature: {bme280['temperature']}")
        self.humidity_label.configure(text=f"Humidity: {bme280['humidity']}")
        self.pressure_label.configure(text=f"Pressure: {bme280['pressure']}")
        self.battery_label.configure(text=f"Battery: {battery}")
        self.distance_label.configure(text=f"Distance: {distance}")

        # データを取得、追加
        self.time_data.append(time)
        acceleration_x = nine_axis['acceleration']['x']
        self.acceleration_x_data.append(acceleration_x)
        acceleration_y = nine_axis['acceleration']['y']
        self.acceleration_y_data.append(acceleration_y)
        acceleration_z = nine_axis['acceleration']['z']
        self.acceleration_z_data.append(acceleration_z)
        angularvelocity_x = nine_axis['angular_velocity']['x']
        self.angularvelocity_x_data.append(angularvelocity_x)
        angularvelocity_y = nine_axis['angular_velocity']['y']
        self.angularvelocity_y_data.append(angularvelocity_y)
        angularvelocity_z = nine_axis['angular_velocity']['z']
        self.angularvelocity_z_data.append(angularvelocity_z)
        Temperature = bme280['temperature']
        self.Temperature_data.append(Temperature)
        Humidity = bme280['humidity']
        self.Humidity_data.append(Humidity)
        Pressure = bme280['pressure']
        self.Pressure_data.append(Pressure)
        Latitude = gps['latitude']
        Longitude = gps['longitude']
        self.coordinates.append((Latitude, Longitude))
        self.battery_data.append((battery))
        self.distance_data.append((distance))

        # グラフにデータをプロット
        self.acx.clear()
        self.acx.plot(self.time_data, self.acceleration_x_data)
        self.acy.clear()
        self.acy.plot(self.time_data, self.acceleration_y_data)
        self.acz.clear()
        self.acz.plot(self.time_data, self.acceleration_z_data)
        self.avx.clear()
        self.avx.plot(self.time_data, self.angularvelocity_x_data)
        self.avy.clear()
        self.avy.plot(self.time_data, self.angularvelocity_y_data)
        self.avz.clear()
        self.avz.plot(self.time_data, self.angularvelocity_z_data)
        self.tem.clear()
        self.tem.plot(self.time_data, self.Temperature_data)
        self.hum.clear()
        self.hum.plot(self.time_data, self.Humidity_data)
        self.pre.clear()
        self.pre.plot(self.time_data, self.Pressure_data)
        # self.bat.clear()
        # self.bat.plot(self.time_data, self.battery_data)
        self.dis.clear()
        self.dis.plot(self.time_data, self.distance_data)
        self.fig_canvas.draw()

        # 地図を作成
        m = folium.Map(location=[0, 0], zoom_start=100)
        # 座標にピンを立てる
        for coord in self.coordinates:
            folium.Marker(coord).add_to(m)
        # 座標を線で結ぶ
        folium.PolyLine(self.coordinates, color='blue').add_to(m)
        # 地図をHTMLファイルに保存
        map_file = "map.html"
        m.save(map_file)

        # Seleniumを使用してHTMLをブラウザで開き、スクリーンショットを取得
        options = webdriver.ChromeOptions()
        options.add_argument('--headless')  # ヘッドレスモードでブラウザを起動
        browser = webdriver.Chrome(options=options)
        wait = WebDriverWait(driver=browser, timeout=10)
        tmpurl = 'file://{path}/{mapfile}'.format(path=os.getcwd(), mapfile=map_file)
        browser.get(tmpurl)
        wait.until(EC.presence_of_all_elements_located)
        browser.save_screenshot("map.png")
        browser.close()
        browser.quit()

    """メッセージ処理"""

    def text_data(self, data):
        time = data.get("time")
        message = data.get("message")
        self.data_text.insert(text=f"Time: {time}")
        self.data_text.insert(text=f"message: {message}")

    """Excelファイル名を生成"""

    def filename(self):
        now = datetime.now()
        self.excel_file_name = "start_" + now.strftime("%Y-%m-%d_%H-%M-%S") + ".xlsx"

        df = pd.DataFrame(columns=[
            "data_type", "time", "gps.latitude", "gps.longitude", "gps.altitude",
            "gps.distance.sample", "gps.distance.goal", "gps.azimuth.sample", "gps.azimuth.goal",
            "nine_axis.acceleration.x", "nine_axis.acceleration.y", "nine_axis.acceleration.z",
            "nine_axis.angular_velocity.x", "nine_axis.angular_velocity.y", "nine_axis.angular_velocity.z",
            "nine_axis.azimuth",
            "bme280.temperature", "bme280.humidity", "bme280.pressure",
            "lps25hb.temperature", "lps25hb.pressure", "lps25hb.altitude",
            "battery", "distance", "camera", "soil_moisture", "message"
        ])
        df.to_excel(self.excel_file_name, index=False)

    """データ保存処理"""

    def save_to_excel(self, data):
        df = pd.read_excel(self.excel_file_name)
        df = pd.concat([df, pd.DataFrame([data])], ignore_index=True)
        print(df)

        # 保存
        with pd.ExcelWriter(self.excel_file_name, engine='openpyxl', mode='a', if_sheet_exists="overlay") as writer:
            df.to_excel(writer, index=False)

        """
        sheet = self.workbook.active
        time = data.get("time")
        gps = data.get("gps")
        nine_axis = data.get("nine_axis")
        bme280 = data.get("bme280")
        battery = data.get("battery")
        distance = data.get("distance")
        lps25hb = data.get("lps25hb")
        # ヘッダーの書き込み
        sheet["A1"] = "Time"
        sheet["B1"] = "Latitude"
        sheet["C1"] = "Longitude"
        sheet["D1"] = "Altitude"
        sheet["E1"] = "Sample Distance"
        sheet["F1"] = "Sample Azimuth"
        sheet["G1"] = "Goal Distance"
        sheet["H1"] = "Goal Azimuth"
        sheet["I1"] = "acc x"
        sheet["J1"] = "acc y"
        sheet["K1"] = "acc z"
        sheet["L1"] = "vel x"
        sheet["M1"] = "vel y"
        sheet["N1"] = "vel z"
        sheet["O1"] = "9 azi"
        sheet["P1"] = "bme tem"
        sheet["Q1"] = "bme hum"
        sheet["R1"] = "bme pre"
        sheet["S1"] = "lps tem"
        sheet["T1"] = "lps pre"
        sheet["U1"] = "lps alt"
        sheet["V1"] = "battery"
        sheet["W1"] = "distance"
        sheet["X1"] = "camera"
        sheet["Y1"] = "soil"
        sheet["Z1"] = "message"

        # JSONデータの書き込み

        sheet.cell(row=self.i, column=1, value=data['time'])
        sheet.cell(row=self.i, column=2, value=data['gps']['latitude'])
        sheet.cell(row=self.i, column=3, value=data['gps']['longitude'])
        sheet.cell(row=self.i, column=4, value=data['gps']['altitude'])
        sheet.cell(row=self.i, column=5, value=data['gps']['distance']['sample'])
        sheet.cell(row=self.i, column=6, value=data['gps']['azimuth']['sample'])
        sheet.cell(row=self.i, column=7, value=data['gps']['distance']['goal'])
        sheet.cell(row=self.i, column=8, value=data['gps']['azimuth']['goal'])
        sheet.cell(row=self.i, column=9, value=data['nine_axis']['acceleration']['x'])
        sheet.cell(row=self.i, column=10, value=data['nine_axis']['acceleration']['y'])
        sheet.cell(row=self.i, column=11, value=data['nine_axis']['acceleration']['z'])
        sheet.cell(row=self.i, column=12, value=data['nine_axis']['angular_velocity']['x'])
        sheet.cell(row=self.i, column=13, value=data['nine_axis']['angular_velocity']['y'])
        sheet.cell(row=self.i, column=14, value=data['nine_axis']['angular_velocity']['z'])
        sheet.cell(row=self.i, column=15, value=data['nine_axis']['azimuth'])
        sheet.cell(row=self.i, column=16, value=data['bme280']['temperature'])
        sheet.cell(row=self.i, column=17, value=data['bme280']['humidity'])
        sheet.cell(row=self.i, column=18, value=data['bme280']['pressure'])
        sheet.cell(row=self.i, column=19, value=data['lps25hb']['temperature'])
        sheet.cell(row=self.i, column=20, value=data['lps25hb']['pressure'])
        sheet.cell(row=self.i, column=21, value=data['lps25hb']['altitude'])
        sheet.cell(row=self.i, column=22, value=data['battery'])
        sheet.cell(row=self.i, column=23, value=data['distance'])
        sheet.cell(row=self.i, column=24, value=data['camera'])
        sheet.cell(row=self.i, column=25, value=data['soil_moisture'])
        sheet.cell(row=self.i, column=26, value=data['message'])
        self.i = self.i + 1
        """

    """写真処理"""

    def picture_data(self, data):
        picture = data.get("camera")
        img_stream = io.BytesIO(picture)
        img = Image.open(img_stream)
        img = img.resize((200, 150))
        photo = ImageTk.PhotoImage(img)
        self.cvs.create_image(200, 150, image=photo, tag="mytest")
        self.cvs.image = photo

    """土壌水分データ処理"""

    def soil_data(self, data):
        soil = data.get("soil_moisture")
        self.soil_label.configure(text=f"soil: {soil}")

    """終了処理"""

    def close(self):
        self.is_serial_connected = False
        if self.serial_port:
            self.serial_port.close()
        self.destroy()

    """スクロールバー"""

    class ScrollableFrame(tk.Frame):
        def __init__(self, parent, *args, **kwargs):
            super().__init__(parent, *args, **kwargs)

            self.canvas = tk.Canvas(self, width=300, height=0, borderwidth=0, highlightthickness=0)
            self.scrollable_frame = tk.Frame(self.canvas)

            self.vsb = tk.Scrollbar(self, orient=tk.VERTICAL, command=self.canvas.yview)
            self.canvas.configure(yscrollcommand=self.vsb.set)

            self.vsb.pack(side=tk.RIGHT, fill=tk.Y)
            self.canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            self.canvas.create_window((0, 0), window=self.scrollable_frame, anchor=tk.NW)

            self.scrollable_frame.bind("<Configure>", self.on_frame_configure)


app = App()
app.mainloop()
